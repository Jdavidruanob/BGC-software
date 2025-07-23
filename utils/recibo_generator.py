# utils/recibo_generator.py

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from datetime import date
from config import format_miles_colombian_int # Asegúrate de que esta función esté aquí o sea accesible

# --- Configuración de rutas (Ajusta según tu estructura de proyecto) ---
# Si utils/recibo_generator.py está en 'tu_proyecto/utils/'
# y la plantilla está en 'tu_proyecto/assets/templates/recibo_template.xlsx'
BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) # Sube dos niveles para llegar a 'tu_proyecto/'
TEMPLATE_REL_PATH = os.path.join("assets", "templates", "recibo_template_op.xlsx") # Ruta relativa a BASE_DIR
TEMPLATE_PATH = os.path.join(BASE_DIR, TEMPLATE_REL_PATH)

OUTPUT_FOLDER_REL_PATH = "recibos_generados"
OUTPUT_FOLDER_PATH = os.path.join(BASE_DIR, OUTPUT_FOLDER_REL_PATH)

# --- Constantes de Celda (Basadas en tus especificaciones) ---
# Se usarán para el acceso inicial, luego se ajustarán dinámicamente
RECIBO_ID_CELL = 'D4'
FECHA_CELL = 'I4'
RECIBI_DE_CELL = 'G6'

# Aportes
APORTE_DATA_START_ROW = 9 # Fila donde empiezan los datos del primer aporte en la plantilla
APORTE_NOMBRE_COL = 'B'
APORTE_SALDO_COL = 'F'
APORTE_MONTO_COL = 'H'
APORTE_NUEVO_SALDO_COL = 'J'
APORTE_TOTAL_CELL_INITIAL = 'H10' # Esta celda se moverá

# Pagos Crédito
CREDITO_DATA_START_ROW = 13 # Fila donde empiezan los datos del primer pago en la plantilla
CREDITO_NOMBRE_COL = 'B'
CREDITO_LETRA_COL = 'F'
CREDITO_CUOTA_COL = 'G'
CREDITO_SDO_CAP_COL = 'H'
CREDITO_AB_CAP_COL = 'I'
CREDITO_INTERES_COL = 'J'
CREDITO_N_SDOCAP_COL = 'K'
CREDITO_TOTAL_CELL_INITIAL = 'H14' # Esta celda se moverá

# Totales Finales
GASTOS_ADMIN_CELL_INITIAL = 'K16' # Esta celda se moverá
TOTAL_GENERAL_CELL_INITIAL = 'K17' # Esta celda se moverá

# Asumo que este es el valor fijo de los gastos de administración.
# Si debe venir de la DB, asegúrate de que tu `on_register` lo pase correctamente.
DEFAULT_GASTOS_ADMIN = 3000 


def generar_recibo_general(
    db_manager, 
    recibo_id: int,
    recibi_de_data: dict, 
    aportes_info: list = None, # Lista de tuplas: (socio_data_dict, monto_aporte_int, saldo_socio_antes, saldo_socio_despues)
    pagos_credito_info: list = None, # Lista de tuplas: (socio_data_dict, letra_id_str, n_cuotas_pagadas, monto_total_pagado_cuotas, detalles_cuotas_list)
    gastos_admin: int = DEFAULT_GASTOS_ADMIN # Usa el valor fijo o el que se pase
):
    """
    Genera un recibo en Excel a partir de una plantilla, rellenando celdas directamente con openpyxl.
    Maneja la inserción de filas para secciones dinámicas.
    """
    if aportes_info is None:
        aportes_info = []
    if pagos_credito_info is None:
        pagos_credito_info = []

    try:
        os.makedirs(OUTPUT_FOLDER_PATH, exist_ok=True)
        file_name = f"Recibo_{recibo_id}_{date.today().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER_PATH, file_name)

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        # --- Reemplazar datos de CABECERA ---
        ws[RECIBO_ID_CELL] = recibo_id
        ws[FECHA_CELL] = date.today().strftime("%d/%m/%Y") # Formato DD/MM/YYYY
        ws[RECIBI_DE_CELL] = f"{recibi_de_data['nombres']} {recibi_de_data['apellidos']}"
        ws[RECIBI_DE_CELL].alignment = Alignment(horizontal='left') # Puedes mantener o ajustar la alineación si es necesario


        # --- PROCESAR APORTES ---
        total_aportes_acumulado = 0
        current_row_for_aportes = APORTE_DATA_START_ROW # Fila inicial para los datos de aportes

        # Determinar cuántas filas se van a insertar para aportes
        num_new_aportes_rows = max(0, len(aportes_info) - 1) # Si hay 1 aporte, 0 filas nuevas; si hay 2, 1 fila nueva, etc.

        if num_new_aportes_rows > 0:
            # Insertar las filas necesarias ANTES de la fila actual de inicio de aportes
            # Esto empuja todo el contenido de abajo (incluyendo la sección de créditos y totales) hacia abajo
            ws.insert_rows(APORTE_DATA_START_ROW + 1, amount=num_new_aportes_rows)
            # Nota: Los estilos de las filas insertadas se copian de la fila anterior.

        # Llenar los datos de aportes
        if aportes_info:
            for i, (socio_data, monto_aporte, saldo_socio_antes, saldo_socio_despues) in enumerate(aportes_info):
                row_to_fill = APORTE_DATA_START_ROW + i

                ws[f'{APORTE_NOMBRE_COL}{row_to_fill}'] = f"{socio_data['nombres']} {socio_data['apellidos']}"
                ws[f'{APORTE_SALDO_COL}{row_to_fill}'] = format_miles_colombian_int(saldo_socio_antes)
                ws[f'{APORTE_MONTO_COL}{row_to_fill}'] = format_miles_colombian_int(monto_aporte)
                ws[f'{APORTE_NUEVO_SALDO_COL}{row_to_fill}'] = format_miles_colombian_int(saldo_socio_despues)
                
                total_aportes_acumulado += monto_aporte
            
            # Ajustar la fila actual para la siguiente sección
            current_row_for_aportes = APORTE_DATA_START_ROW + len(aportes_info) - 1 # Última fila donde se escribió un aporte
        else: # Si no hay aportes, limpiar la fila de ejemplo (APORTE_DATA_START_ROW)
            for col_letter in [APORTE_NOMBRE_COL, APORTE_SALDO_COL, APORTE_MONTO_COL, APORTE_NUEVO_SALDO_COL]:
                ws[f'{col_letter}{APORTE_DATA_START_ROW}'] = "" # Limpiar el contenido de la primera fila
            ws[f'{APORTE_NOMBRE_COL}{APORTE_DATA_START_ROW}'] = "N/A" # Opcional: indicar que no hay aportes
            current_row_for_aportes = APORTE_DATA_START_ROW # La fila de aportes "termina" en su fila original


        # Calcular la nueva posición de la celda de Total Aportes
        # La fila H10 original se habrá movido por las filas insertadas
        total_aportes_row_offset = (APORTE_DATA_START_ROW + len(aportes_info)) - int(APORTE_TOTAL_CELL_INITIAL[1:]) # (fila donde termina el último aporte) - (fila original del total de aportes)
        new_total_aportes_row = int(APORTE_TOTAL_CELL_INITIAL[1:]) + num_new_aportes_rows 
        ws[f'{APORTE_TOTAL_CELL_INITIAL[0]}{new_total_aportes_row}'] = format_miles_colombian_int(total_aportes_acumulado)


        # --- PROCESAR PAGOS DE CRÉDITO ---
        # La posición de inicio de los créditos se ve afectada por las filas insertadas en aportes
        current_credito_section_start_row = CREDITO_DATA_START_ROW + num_new_aportes_rows
        
        total_creditos_acumulado = 0
        current_row_for_creditos_data = current_credito_section_start_row

        # Determinar cuántas filas se van a insertar para créditos
        # Cada pago de crédito se representa en una fila
        num_new_creditos_rows = max(0, len(pagos_credito_info) - 1) 

        if num_new_creditos_rows > 0:
            ws.insert_rows(current_credito_section_start_row + 1, amount=num_new_creditos_rows)
            

        # Llenar los datos de pagos de crédito
        if pagos_credito_info:
            for i, (socio_data, letra_id_str, n_cuotas_pagadas_group, monto_total_pagado_group, detalles_cuotas_list) in enumerate(pagos_credito_info):
                # Para cada elemento en detalles_cuotas_list, si es un pago de varias cuotas
                for j, detalle_cuota in enumerate(detalles_cuotas_list):
                    # Calcula la fila donde se escribirá esta cuota
                    # (current_credito_section_start_row + i_del_grupo + j_de_la_cuota_en_el_grupo)
                    row_to_fill_credito = current_credito_section_start_row + i # Aquí i es el índice del pago_credito_info
                    
                    # Si no es la primera cuota del primer crédito, debemos insertar una nueva fila
                    if (i > 0 or j > 0):
                        ws.insert_rows(row_to_fill_credito + 1) # Insertar justo debajo de la última fila rellenada
                        current_row_for_creditos_data += 1 # Actualiza el puntero de la fila actual

                    # Rellena las celdas para esta cuota específica
                    ws[f'{CREDITO_NOMBRE_COL}{current_row_for_creditos_data}'] = f"{socio_data['nombres']} {socio_data['apellidos']}"
                    ws[f'{CREDITO_LETRA_COL}{current_row_for_creditos_data}'] = detalle_cuota['letra_id']
                    
                    # Obtener el total de cuotas del crédito
                    total_cuotas_credito = db_manager.get_total_cuotas_credito(detalle_cuota['letra_id']) 
                    # El formato de cuota es "1-3/36" o "1/36"
                    # Si 'nro_cuota' es la cuota actual, y queremos mostrar un rango
                    # Esto requiere más lógica si quieres "1-3/36". Por ahora, solo la cuota actual.
                    cuota_display = f"{detalle_cuota['nro_cuota']} / {total_cuotas_credito}"
                    
                    ws[f'{CREDITO_CUOTA_COL}{current_row_for_creditos_data}'] = cuota_display
                    ws[f'{CREDITO_SDO_CAP_COL}{current_row_for_creditos_data}'] = format_miles_colombian_int(detalle_cuota['saldo_capital_antes_pago']) 
                    ws[f'{CREDITO_AB_CAP_COL}{current_row_for_creditos_data}'] = format_miles_colombian_int(detalle_cuota['valor_cuota'])
                    ws[f'{CREDITO_INTERES_COL}{current_row_for_creditos_data}'] = format_miles_colombian_int(detalle_cuota['interes_mes'])
                    ws[f'{CREDITO_N_SDOCAP_COL}{current_row_for_creditos_data}'] = format_miles_colombian_int(detalle_cuota['saldo_capital_despues_pago'])
                    
                    total_creditos_acumulado += (detalle_cuota['valor_cuota'] + detalle_cuota['interes_mes'])
            
            # Ajustar la fila actual para la siguiente sección
            # current_row_for_creditos_data ya está actualizada por el bucle interno
        else: # Si no hay pagos de crédito, limpiar la fila de ejemplo (CREDITO_DATA_START_ROW + offset)
            row_to_clear_creditos = current_credito_section_start_row 
            for col_letter in [CREDITO_NOMBRE_COL, CREDITO_LETRA_COL, CREDITO_CUOTA_COL, 
                               CREDITO_SDO_CAP_COL, CREDITO_AB_CAP_COL, CREDITO_INTERES_COL, CREDITO_N_SDOCAP_COL]:
                ws[f'{col_letter}{row_to_clear_creditos}'] = ""
            ws[f'{CREDITO_NOMBRE_COL}{row_to_clear_creditos}'] = "N/A" # Opcional: indicar que no hay pagos
            current_row_for_creditos_data = row_to_clear_creditos # La sección termina en su fila original (después del offset)


        # Calcular la nueva posición de la celda de Total Créditos
        # La fila H14 original se habrá movido por las filas insertadas de aportes y créditos
        new_total_creditos_row = int(CREDITO_TOTAL_CELL_INITIAL[1:]) + num_new_aportes_rows + num_new_creditos_rows 
        ws[f'{CREDITO_TOTAL_CELL_INITIAL[0]}{new_total_creditos_row}'] = format_miles_colombian_int(total_creditos_acumulado)

        
        # --- GASTOS ADMINISTRACIÓN y TOTAL GENERAL ---
        # Las posiciones originales K16 y K17 también se habrán movido
        new_gastos_admin_row = int(GASTOS_ADMIN_CELL_INITIAL[1:]) + num_new_aportes_rows + num_new_creditos_rows
        new_total_general_row = int(TOTAL_GENERAL_CELL_INITIAL[1:]) + num_new_aportes_rows + num_new_creditos_rows

        ws[f'{GASTOS_ADMIN_CELL_INITIAL[0]}{new_gastos_admin_row}'] = format_miles_colombian_int(gastos_admin)
        
        total_general = total_aportes_acumulado + total_creditos_acumulado + gastos_admin
        ws[f'{TOTAL_GENERAL_CELL_INITIAL[0]}{new_total_general_row}'] = format_miles_colombian_int(total_general)

        # --- Guardar el recibo ---
        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"Error al generar recibo: {e}")
        import traceback
        traceback.print_exc() # Imprime el stack trace completo para depuración
        return None