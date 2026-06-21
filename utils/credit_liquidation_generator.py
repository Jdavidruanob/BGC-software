# utils/credit_liquidation_generator.py

import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, datetime
from dateutil.relativedelta import relativedelta

from config import (
    format_miles_colombian_int, 
    ASSETS_DIR, LIQUIDACIONES_OUTPUT_DIR
)

# --- Rutas y Constantes ---
TEMPLATE_LIQUIDACION_REL_PATH = os.path.join("templates", "recibo_template_liquidacion.xlsx")
TEMPLATE_LIQUIDACION_PATH = os.path.join(ASSETS_DIR, TEMPLATE_LIQUIDACION_REL_PATH)

# La carpeta de salida para las liquidaciones
LIQUIDACIONES_OUTPUT_DIR = LIQUIDACIONES_OUTPUT_DIR

# Constantes de Celda para el Encabezado (según tu descripción)
LETRA_CREDITO_CELL = 'B7'
CAPITAL_CREDITO_CELL = 'F7'
SOCIOS_PARTICIPANTES_CELL = 'B9'
# CAMBIO 1: Ahora usaremos esta celda para la FECHA DE PRIMER VENCIMIENTO
FECHA_CREACION_CREDITO = 'A12' 
NO_CUOTAS_CELL = 'B12'
VALOR_CUOTA_ESTIMADA_CELL = 'C12'
INTERES_CREDITO_CELL = 'D12'
CAPITAL_PRESTADO_REPETIDO_CELL = 'F12' # Mismo que CAPITAL_CREDITO_CELL

# Configuración de la tabla de liquidación
TABLA_START_ROW = 14
TABLA_HEADERS = [
    "Fecha", "Cuota", "Valor Cuota", "Intereses",
    "Total Mensual", "Saldo Capital", "Fecha Pago"
]
# Columnas de la A a la G para la tabla
TABLA_COLUMNS = ['A', 'B', 'C', 'D', 'E', 'F', 'G']

# Estilos básicos para la tabla
BORDER_STYLE = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000')
)
HEADER_FONT = Font(bold=True)
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

# --- Función para Generar la Liquidación del Crédito en Excel ---
def generar_liquidacion_credito(
    credit_data: dict,  # Diccionario con 'letra', 'capital', 'interes', 'no_cuotas', 'fecha_inicio'
    socios_list: list   # Lista de diccionarios de socios, cada uno con 'nombres' y 'apellidos'
):
    """
    Genera un archivo Excel con la liquidación detallada de un crédito.

    Args:
        credit_data (dict): Datos del crédito (letra, capital, interes, no_cuotas, fecha_inicio).
        socios_list (list): Lista de diccionarios de socios participantes, cada uno con 'nombres' y 'apellidos'.
    """
    try:
        # Asegúrate de que la carpeta de salida exista
        os.makedirs(LIQUIDACIONES_OUTPUT_DIR, exist_ok=True)

        file_name = f"Liquidacion_letra_{credit_data['letra']}_{date.today().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(LIQUIDACIONES_OUTPUT_DIR, file_name)

        wb = load_workbook(TEMPLATE_LIQUIDACION_PATH)
        ws = wb.active

        # --- Llenar el Encabezado del Crédito ---
        ws[LETRA_CREDITO_CELL] = credit_data['letra']
        ws[CAPITAL_CREDITO_CELL] = format_miles_colombian_int(credit_data['capital'])
        
        # Socios participantes (en mayúsculas, separados por Y/O)
        socios_nombres_completos = [f"{s['nombres']} {s['apellidos']}".upper() for s in socios_list]
        ws[SOCIOS_PARTICIPANTES_CELL] = ", Y/O ".join(socios_nombres_completos)

        fecha_inicio_dt = datetime.strptime(credit_data["fecha_inicio"][:10], "%Y-%m-%d")
        
        # CAMBIO 1: Calcular la fecha de la primera cuota para el "Vencimiento"
        ws[FECHA_CREACION_CREDITO] = fecha_inicio_dt.strftime("%Y-%m-%d")

        ws[NO_CUOTAS_CELL] = credit_data['no_cuotas']
        ws[INTERES_CREDITO_CELL] = f"{credit_data['interes'] * 100:.2f}%"
        ws[CAPITAL_PRESTADO_REPETIDO_CELL] = format_miles_colombian_int(credit_data['capital'])

        # --- Calcular la tabla de liquidación (lógica de CreditLiquidationPage) ---
        capital = credit_data["capital"]
        interes = credit_data["interes"]
        cuotas = credit_data["no_cuotas"]
        
        cuota_base = None
        cuota_final = None

        # Buscar el mejor redondeo que cumpla condiciones
        for redondeo in [10000, 9000, 8000, 7000, 6000, 5000, 2000, 1000]:
            posible_cuota = round((capital / cuotas) / redondeo) * redondeo
            total_normales = posible_cuota * (cuotas - 1)
            ultima_cuota = capital - total_normales

            if 10000 <= ultima_cuota >= 0 and ultima_cuota <= posible_cuota * 1.5: # Asegurar que última cuota no sea negativa
                cuota_base = posible_cuota
                cuota_final = ultima_cuota
                break

        if cuota_base is None:
            # Último recurso: sin redondear
            cuota_base = capital // cuotas
            cuota_final = capital - cuota_base * (cuotas - 1)
            # Ajustar si la última cuota queda muy pequeña o grande
            if cuota_final < 10000 and cuotas > 1: # Reducir cuota base si la última es muy pequeña
                cuota_base -= 1000
                cuota_final = capital - cuota_base * (cuotas - 1)
            elif cuota_final > cuota_base * 1.5 and cuotas > 1: # Aumentar cuota base si la última es muy grande
                 cuota_base += 1000
                 cuota_final = capital - cuota_base * (cuotas - 1)

        # Asigna el valor de la cuota base al encabezado
        ws[VALOR_CUOTA_ESTIMADA_CELL] = format_miles_colombian_int(cuota_base)


        saldo = capital
        current_row = TABLA_START_ROW # Fila donde inician los encabezados de la tabla

        # Escribir encabezados de la tabla
        for col_idx, header_text in enumerate(TABLA_HEADERS):
            cell = ws[f"{TABLA_COLUMNS[col_idx]}{current_row}"]
            cell.value = header_text
            cell.font = HEADER_FONT
            cell.alignment = CENTER_ALIGNMENT
            cell.border = BORDER_STYLE
            
        current_row += 1 # Pasa a la siguiente fila para los datos

        # Llenar la tabla con los datos de la liquidación
        for i in range(cuotas):
            nro_cuota = i + 1
            fecha = fecha_inicio_dt + relativedelta(months=+nro_cuota) # Fecha de vencimiento de cada cuota
            cuota_valor = cuota_final if i == cuotas - 1 else cuota_base
            intereses = round(saldo * interes)
            cuota_mensual = cuota_valor + intereses
            saldo -= cuota_valor

            row_data = [
                fecha.strftime("%Y-%m-%d"),
                str(nro_cuota),
                format_miles_colombian_int(cuota_valor),
                format_miles_colombian_int(intereses),
                format_miles_colombian_int(cuota_mensual),
                format_miles_colombian_int(max(0, saldo)), # Saldo no debe ser negativo
                "" # Fecha Pago (vacío para una liquidación nueva)
            ]

            for col_idx, value in enumerate(row_data):
                cell = ws[f"{TABLA_COLUMNS[col_idx]}{current_row}"]
                cell.value = value
                cell.alignment = CENTER_ALIGNMENT
                cell.border = BORDER_STYLE
            
            current_row += 1 # Pasa a la siguiente fila para la próxima cuota

        # Ajustar ancho de columnas (opcional, para mejor visualización)
        for col_idx in range(len(TABLA_COLUMNS)):
            ws.column_dimensions[TABLA_COLUMNS[col_idx]].width = 15 


        # --- CAMBIO 2: Añadir la línea del tesorero ---
        signature_row = current_row + 2 # Dos filas debajo de la última fila de la tabla
        
        # Combinar celdas de la A a la G en la fila de la firma
        ws.merge_cells(start_row=signature_row, start_column=1, end_row=signature_row, end_column=7)
        
        # Escribir el texto en la celda combinada (siempre a la superior izquierda)
        signature_cell = ws[f"A{signature_row}"]
        signature_cell.value = "Tesorero: ALVARO L. BURBANO GARCIA"
        signature_cell.alignment = CENTER_ALIGNMENT
        signature_cell.font = Font(bold=True, size=12) # Opcional: negrita y tamaño

        # --- Guardar el archivo ---
        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"❌ Error al generar liquidación de crédito: {e}")
        import traceback
        traceback.print_exc()
        return None