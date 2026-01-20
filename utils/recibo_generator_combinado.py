import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment 
from datetime import date
from config import (
    format_miles_colombian_int, 
    format_full_name_for_excel, 
    ASSETS_DIR, RECIBOS_OUTPUT_DIR
)

# Carpeta de Salida
OUTPUT_FOLDER_PATH = RECIBOS_OUTPUT_DIR

# --- Constantes Generales ---
RECIBO_ID_CELL = 'D4'
FECHA_CELL = 'I4'
RECIBI_DE_CELL = 'G6'

# --- CONSTANTES DE COLUMNAS (Letras) ---
# Sección Aportes
APORTE_NOMBRE_COL = 'B'
APORTE_SALDO_COL = 'F' 
APORTE_MONTO_COL = 'H' 
APORTE_NUEVO_SALDO_COL = 'J' 
COL_TOTAL_APORTES = 'H'

# Sección Créditos
CREDITO_NOMBRE_COL = 'B' 
CREDITO_LETRA_COL = 'F'
CREDITO_CUOTA_COL = 'G'
CREDITO_SDO_CAP_COL = 'H'
CREDITO_AB_CAP_COL = 'I'
CREDITO_INTERES_COL = 'J' 
CREDITO_N_SDOCAP_COL = 'K'
COL_TOTAL_CREDITOS = 'H'

# Totales Finales
COL_RESUMEN_VALORES = 'K'

GASTO_POR_APORTE = 3000
MAX_FILAS_PERMITIDAS = 6 

def generar_recibo_combinado(
    db_manager, 
    recibo_id: int,
    recibi_de_data: dict, 
    aportes_info: list = None, 
    pagos_credito_info: list = None,
    num_aportes_cobrables: int = None
):
    """
    Genera un recibo combinado seleccionando la plantilla exacta según la matriz:
    recibo_template_combinado{X}_{Y}.xlsx
    Donde X = num_aportes, Y = num_pagos.
    Incluye lógica de Mora.
    """
    if aportes_info is None: aportes_info = []
    if pagos_credito_info is None: pagos_credito_info = []

    # 1. Determinar dimensiones de la matriz (X, Y)
    num_aportes = len(aportes_info)
    num_pagos = len(pagos_credito_info)

    if num_aportes == 0 and num_pagos == 0:
        return None
    
    if num_aportes > MAX_FILAS_PERMITIDAS:
        aportes_info = aportes_info[:MAX_FILAS_PERMITIDAS]
        num_aportes = MAX_FILAS_PERMITIDAS
        
    if num_pagos > MAX_FILAS_PERMITIDAS:
        pagos_credito_info = pagos_credito_info[:MAX_FILAS_PERMITIDAS]
        num_pagos = MAX_FILAS_PERMITIDAS

    try:
        os.makedirs(OUTPUT_FOLDER_PATH, exist_ok=True)
        file_name = f"Recibo_{recibo_id}_{date.today().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER_PATH, file_name)

        # 2. SELECCIÓN DE PLANTILLA
        template_name = f"recibo_template_combinado{num_aportes}_{num_pagos}.xlsx"
        template_rel_path = os.path.join("templates", "recibo_template_combinado", template_name)
        template_abs_path = os.path.join(ASSETS_DIR, template_rel_path)

        if not os.path.exists(template_abs_path):
            print(f"❌ Error CRÍTICO: No se encontró la plantilla {template_name}")
            return None

        wb = load_workbook(template_abs_path) 
        ws = wb.active

        # --- CABECERA ---
        ws[RECIBO_ID_CELL] = recibo_id
        ws[FECHA_CELL] = date.today().strftime("%d/%m/%Y")
        recibi_de_full_name = f"{recibi_de_data['nombres']} {recibi_de_data['apellidos']}".upper()
        ws[RECIBI_DE_CELL] = recibi_de_full_name
        ws[RECIBI_DE_CELL].alignment = Alignment(horizontal='center') 

        # ==========================================
        # 3. SECCIÓN APORTES
        # ==========================================
        start_row_aportes = 9
        total_aportes_monto = 0 

        for i in range(num_aportes):
            row = start_row_aportes + i
            detalle = aportes_info[i]
            
            socio = detalle['socio_data']
            nombre = format_full_name_for_excel(socio['nombres'], socio['apellidos'])
            
            ws[f'{APORTE_NOMBRE_COL}{row}'] = nombre
            ws[f'{APORTE_SALDO_COL}{row}'] = format_miles_colombian_int(detalle['saldo_anterior']) 
            ws[f'{APORTE_MONTO_COL}{row}'] = format_miles_colombian_int(detalle['monto'])
            ws[f'{APORTE_NUEVO_SALDO_COL}{row}'] = format_miles_colombian_int(detalle['saldo_nuevo']) 
            
            total_aportes_monto += detalle['monto']

        row_total_aportes = start_row_aportes + num_aportes
        ws[f'{COL_TOTAL_APORTES}{row_total_aportes}'] = format_miles_colombian_int(total_aportes_monto)

        # ==========================================
        # 4. SECCIÓN CRÉDITOS
        # ==========================================
        start_row_creditos = row_total_aportes + 3
        
        total_creditos_monto = 0 
        total_mora_recibo = 0 # Acumulador Mora
        cuotas_cache = {}

        for i in range(num_pagos):
            row = start_row_creditos + i
            detalle = pagos_credito_info[i]
            
            socio = detalle['socio_data'] 
            letra_id = detalle['letra_id']
            nombre = format_full_name_for_excel(socio['nombres'], socio['apellidos'])
            
            # Sumar mora al total
            total_mora_recibo += detalle.get('mora_consolidada', 0)

            # Datos básicos
            ws[f'{CREDITO_NOMBRE_COL}{row}'] = nombre
            ws[f'{CREDITO_LETRA_COL}{row}'] = letra_id
            
            # Cuota / Abono Logic
            n_start = detalle['nro_cuotas_pagadas_start']
            n_end = detalle['nro_cuotas_pagadas_end']
            
            es_abono = isinstance(n_start, str) and "ABONO" in n_start
            
            if es_abono:
                cuota_txt = "ABONO"
            else:
                if letra_id not in cuotas_cache:
                    cuotas_cache[letra_id] = db_manager.get_total_cuotas_credito(letra_id)
                total_c = cuotas_cache[letra_id]
                
                if n_start == n_end:
                    cuota_txt = f"{n_start}/{total_c}"
                else:
                    cuota_txt = f"{n_start}-{n_end}/{total_c}"
            
            ws[f'{CREDITO_CUOTA_COL}{row}'] = cuota_txt
            
            # Valores Monetarios
            ws[f'{CREDITO_SDO_CAP_COL}{row}'] = format_miles_colombian_int(detalle['saldo_capital_antes_pago']) 
            ws[f'{CREDITO_AB_CAP_COL}{row}'] = format_miles_colombian_int(detalle['valor_capital_consolidado'])
            ws[f'{CREDITO_INTERES_COL}{row}'] = format_miles_colombian_int(detalle['interes_consolidado'])
            ws[f'{CREDITO_N_SDOCAP_COL}{row}'] = format_miles_colombian_int(detalle['saldo_capital_despues_pago'])
            
            total_creditos_monto += (detalle['valor_capital_consolidado'] + detalle['interes_consolidado'])

        row_total_creditos = start_row_creditos + num_pagos
        ws[f'{COL_TOTAL_CREDITOS}{row_total_creditos}'] = format_miles_colombian_int(total_creditos_monto)

        # ==========================================
        # 5. PIE DE PÁGINA (Totales Finales)
        # ==========================================
        row_admin = row_total_creditos + 2
        row_mora = row_admin + 1
        row_general = row_mora + 1

        # 1. Gastos Admin
        if num_aportes_cobrables is not None:
            cant_cobrar = num_aportes_cobrables
        else:
            cant_cobrar = num_aportes
            
        val_admin = GASTO_POR_APORTE * cant_cobrar
        ws[f'{COL_RESUMEN_VALORES}{row_admin}'] = format_miles_colombian_int(val_admin) 
        
        # 2. Interés Mora
        ws[f'{COL_RESUMEN_VALORES}{row_mora}'] = format_miles_colombian_int(total_mora_recibo)

        # 3. Total General
        val_general = total_aportes_monto + total_creditos_monto + val_admin + total_mora_recibo
        
        ws[f'{COL_RESUMEN_VALORES}{row_general}'] = format_miles_colombian_int(val_general) 
    
        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"Error al generar recibo combinado matricial: {e}")
        import traceback
        traceback.print_exc() 
        return None