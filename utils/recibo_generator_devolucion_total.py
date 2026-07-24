import os
from openpyxl import load_workbook
from datetime import date
from config import (
    format_miles_colombian_int,
    ASSETS_DIR, RECIBOS_OUTPUT_DIR, get_hoy,
)

# --- Rutas y Constantes ---
TEMPLATE_REL_PATH = os.path.join("templates", "recibo_template_devolucion_total.xlsx")
TEMPLATE_PATH = os.path.join(ASSETS_DIR, TEMPLATE_REL_PATH)

OUTPUT_FOLDER_PATH = RECIBOS_OUTPUT_DIR

# --- Celdas del template 'devolucion-total.xlsx' ---
RECIBO_ID_CELL = 'B5'   # Numero de recibo
VALOR_CELL = 'F5'       # Valor devuelto (al lado de "Por:")
FECHA_CELL = 'B7'       # Fecha
SOCIO_CELL = 'B9'       # Nombre completo del socio


def generar_recibo_devolucion_total(recibo_id: int, socio_data: dict, valor: int):
    """Genera el recibo de devolución/retiro TOTAL.

    Args:
        recibo_id: ID único del recibo.
        socio_data: dict con 'nombres' y 'apellidos'.
        valor: monto total devuelto al socio.
    """
    try:
        os.makedirs(OUTPUT_FOLDER_PATH, exist_ok=True)

        file_name = f"Devolucion_total_{recibo_id}_{get_hoy().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER_PATH, file_name)

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        ws[RECIBO_ID_CELL] = recibo_id
        ws[VALOR_CELL] = format_miles_colombian_int(valor)
        ws[FECHA_CELL] = get_hoy().strftime("%d/%m/%Y")
        ws[SOCIO_CELL] = f"{socio_data['nombres']} {socio_data['apellidos']}".upper()

        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"Error al generar recibo de devolución total: {e}")
        import traceback
        traceback.print_exc()
        return None
