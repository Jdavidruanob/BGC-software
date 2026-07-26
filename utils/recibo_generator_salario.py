import os

from openpyxl import load_workbook

from config import (
    ASSETS_DIR, RECIBOS_OUTPUT_DIR, format_miles_colombian_int, get_hoy,
)

# --- Rutas y Constantes ---
TEMPLATE_REL_PATH = os.path.join("templates", "recibo_template_salario.xlsx")
TEMPLATE_PATH = os.path.join(ASSETS_DIR, TEMPLATE_REL_PATH)

OUTPUT_FOLDER_PATH = RECIBOS_OUTPUT_DIR

# --- Celdas del template 'recibo_template_salario.xlsx' ---
# Mismas celdas que usa la API (coop_api.recibos.generador.generar_xlsx_salario):
# las combinadas se llenan desde su celda superior-izquierda.
RECIBO_ID_CELL = "B4"   # Número de recibo
VALOR_CELL = "G4"       # Valor del salario
FECHA_CELL = "D6"       # Fecha
MES_CELL = "C12"        # Mes en palabra, p.ej. "Junio"

MESES = [
    "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
    "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre",
]


def nombre_mes(fecha=None) -> str:
    """Mes en palabra. Se calcula aquí y no con `strftime('%B')` porque el
    nombre del mes depende del locale del sistema, que en Windows no está
    garantizado en español."""
    fecha = fecha or get_hoy()
    return MESES[fecha.month - 1]


def generar_recibo_salario(recibo_id: int, valor: int, mes: str = None):
    """Genera el recibo del salario del administrador.

    Args:
        recibo_id: ID único del recibo.
        valor: monto pagado.
        mes: mes al que corresponde el salario. Si no se pasa, el mes actual.
    """
    try:
        os.makedirs(OUTPUT_FOLDER_PATH, exist_ok=True)

        file_name = f"Recibo_{recibo_id}_{get_hoy().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER_PATH, file_name)

        wb = load_workbook(TEMPLATE_PATH)
        ws = wb.active

        ws[RECIBO_ID_CELL] = recibo_id
        ws[VALOR_CELL] = format_miles_colombian_int(valor)
        ws[FECHA_CELL] = get_hoy().strftime("%d/%m/%Y")
        ws[MES_CELL] = mes or nombre_mes()

        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"Error al generar recibo de salario: {e}")
        import traceback
        traceback.print_exc()
        return None
