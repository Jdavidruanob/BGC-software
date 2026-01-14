import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment 
from datetime import date
from config import (
    format_miles_colombian_int, 
    format_full_name_for_excel, 
    ASSETS_DIR, RECIBOS_OUTPUT_DIR
)

# Carpeta de Salida
OUTPUT_FOLDER_PATH = RECIBOS_OUTPUT_DIR

# --- Constantes Generales ---
RECIBO_ID_CELL = 'D4'
FECHA_CELL = 'I4'
RECIBI_DE_CELL = 'G6'

# La fila donde SIEMPRE comienzan los datos (en todas las plantillas 1-6)
CREDITO_DATA_START_ROW = 9 

# Columnas (Letras)
CREDITO_NOMBRE_COL = 'B'
CREDITO_LETRA_COL = 'F'
CREDITO_CUOTA_COL = 'G' 
CREDITO_SDO_CAP_COL = 'H' 
CREDITO_AB_CAP_COL = 'I' 
CREDITO_INTERES_COL = 'J' 
CREDITO_N_SDOCAP_COL = 'K' 

# Columnas para totales (Las filas son dinámicas)
COL_TOTAL_TABLA = 'H'      # Total de la tabla (Abono Cap + Interés)
COL_RESUMEN_VALORES = 'K'  # Columna donde van los valores del resumen final

MAX_PAGOS_PERMITIDOS = 6  # Límite de filas

def generar_recibo_solo_pagos(
    db_manager, 
    recibo_id: int,
    recibi_de_data: dict, 
    pagos_credito_info: list = None
):
    """
    Genera un recibo de pagos de crédito seleccionando la plantilla exacta (1 a 6).
    Incluye lógica para fila de 'Interés por Mora'.
    """
    if pagos_credito_info is None:
        pagos_credito_info = []

    # 1. Determinar cuántos pagos hay para elegir la plantilla
    num_pagos = len(pagos_credito_info)

    if num_pagos == 0:
        print("Error: No hay pagos para generar recibo.")
        return None
    
    if num_pagos > MAX_PAGOS_PERMITIDOS:
        print(f"Advertencia: Hay {num_pagos} pagos. Se truncará a {MAX_PAGOS_PERMITIDOS}.")
        pagos_credito_info = pagos_credito_info[:MAX_PAGOS_PERMITIDOS]
        num_pagos = MAX_PAGOS_PERMITIDOS

    try:
        os.makedirs(OUTPUT_FOLDER_PATH, exist_ok=True)
        file_name = f"Recibo_{recibo_id}_{date.today().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER_PATH, file_name)

        # 2. SELECCIÓN DE PLANTILLA SEGÚN CANTIDAD
        # Carpeta: assets/templates/recibo_template_pago/
        # Archivo: recibo_template_pagoX.xlsx
        template_name = f"recibo_template_pago{num_pagos}.xlsx"
        template_rel_path = os.path.join("templates", "recibo_template_pago", template_name)
        template_abs_path = os.path.join(ASSETS_DIR, template_rel_path)

        if not os.path.exists(template_abs_path):
            print(f"❌ Error CRÍTICO: No se encontró la plantilla {template_name}")
            return None

        wb = load_workbook(template_abs_path) 
        ws = wb.active

        # --- CABECERA ---
        ws[RECIBO_ID_CELL] = recibo_id
        ws[FECHA_CELL] = date.today().strftime("%d/%m/%Y")
        
        recibi_de_full_name = f"{recibi_de_data['nombres']} {recibi_de_data['apellidos']}".upper()
        ws[RECIBI_DE_CELL] = recibi_de_full_name
        ws[RECIBI_DE_CELL].alignment = Alignment(horizontal='center') 

        # --- CUERPO (Llenar filas exactas) ---
        total_acumulado_capital_interes = 0
        total_mora_recibo = 0 # Variable para acumular la mora de todas las filas
        
        # Cache para almacenar el número total de cuotas por letra
        cuotas_info_cache = {}

        for i in range(num_pagos):
            row_to_fill = CREDITO_DATA_START_ROW + i
            
            detalle_consolidado = pagos_credito_info[i]
            socio_data = detalle_consolidado['socio_data'] 
            letra_id = detalle_consolidado['letra_id']
            nro_cuotas_start = detalle_consolidado['nro_cuotas_pagadas_start']
            nro_cuotas_end = detalle_consolidado['nro_cuotas_pagadas_end']

            # Acumular Mora (Si existe en el diccionario)
            mora_fila = detalle_consolidado.get('mora_consolidada', 0)
            total_mora_recibo += mora_fila

            formatted_socio_name = format_full_name_for_excel(
                socio_data['nombres'], 
                socio_data['apellidos'], 
                max_length=24
            )
            ws[f'{CREDITO_NOMBRE_COL}{row_to_fill}'] = formatted_socio_name
            ws[f'{CREDITO_LETRA_COL}{row_to_fill}'] = letra_id

            # --- Lógica Visual de Cuotas vs Abono ---
            es_abono = False
            if isinstance(nro_cuotas_start, str) and "ABONO" in nro_cuotas_start:
                es_abono = True

            if es_abono:
                cuota_display = "NA" # O "ABONO"
            else:
                if letra_id not in cuotas_info_cache:
                    total_cuotas_credito = db_manager.get_total_cuotas_credito(letra_id)
                    cuotas_info_cache[letra_id] = total_cuotas_credito
                
                if nro_cuotas_start == nro_cuotas_end:
                    cuota_display = f"{nro_cuotas_start}/{cuotas_info_cache[letra_id]}"
                else:
                    cuota_display = f"{nro_cuotas_start}-{nro_cuotas_end}/{cuotas_info_cache[letra_id]}"
            
            ws[f'{CREDITO_CUOTA_COL}{row_to_fill}'] = cuota_display
            
            ws[f'{CREDITO_SDO_CAP_COL}{row_to_fill}'] = format_miles_colombian_int(detalle_consolidado['saldo_capital_antes_pago']) 
            ws[f'{CREDITO_AB_CAP_COL}{row_to_fill}'] = format_miles_colombian_int(detalle_consolidado['valor_capital_consolidado'])
            ws[f'{CREDITO_INTERES_COL}{row_to_fill}'] = format_miles_colombian_int(detalle_consolidado['interes_consolidado'])
            ws[f'{CREDITO_N_SDOCAP_COL}{row_to_fill}'] = format_miles_colombian_int(detalle_consolidado['saldo_capital_despues_pago'])
            
            total_acumulado_capital_interes += (detalle_consolidado['valor_capital_consolidado'] + detalle_consolidado['interes_consolidado'])

        # --- PIE DE PÁGINA (Cálculo dinámico con nueva fila MORA) ---
        
        # Fila Total Tabla (inmediatamente después de los datos)
        row_total_tabla = CREDITO_DATA_START_ROW + num_pagos
        
        # Geometría asumida:
        # Total Tabla -> +2 filas -> Gastos Admin
        # Gastos Admin -> +1 fila -> Mora (NUEVO)
        # Mora -> +1 fila -> Total General
        
        row_gastos_admin = row_total_tabla + 2
        row_mora = row_gastos_admin + 1        # <--- NUEVA FILA
        row_total_general = row_mora + 1       # <--- DESPLAZADO

        # 1. Total de la Tabla (Capital + Interés corriente)
        ws[f'{COL_TOTAL_TABLA}{row_total_tabla}'] = format_miles_colombian_int(total_acumulado_capital_interes)

        # 2. Gastos Administración (0 por defecto en solo pagos, a menos que cambies la lógica)
        val_admin = 0
        ws[f'{COL_RESUMEN_VALORES}{row_gastos_admin}'] = format_miles_colombian_int(val_admin)
        
        # 3. Interés por Mora (VALOR CALCULADO)
        ws[f'{COL_RESUMEN_VALORES}{row_mora}'] = format_miles_colombian_int(total_mora_recibo)

        # 4. Total General
        total_general = total_acumulado_capital_interes + val_admin + total_mora_recibo
        ws[f'{COL_RESUMEN_VALORES}{row_total_general}'] = format_miles_colombian_int(total_general)

        # --- Guardar ---
        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"Error al generar recibo exacto de pagos: {e}")
        import traceback
        traceback.print_exc() 
        return None