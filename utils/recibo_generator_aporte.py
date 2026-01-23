import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment 
from datetime import date
from config import (
    format_miles_colombian_int, 
    format_full_name_for_excel, 
    ASSETS_DIR, RECIBOS_OUTPUT_DIR, get_hoy
)

# Renombrado y ajustado para tu estructura deseada
OUTPUT_FOLDER_PATH = RECIBOS_OUTPUT_DIR 

# --- Constantes Generales ---
RECIBO_ID_CELL = 'D4'
FECHA_CELL = 'I4'
RECIBI_DE_CELL = 'G6'

# La fila donde SIEMPRE comienzan los datos (en todas las plantillas 1-6)
APORTE_DATA_START_ROW = 9 

# Columnas (Letras)
APORTE_NOMBRE_COL = 'B'
APORTE_SALDO_COL = 'F'
APORTE_MONTO_COL = 'H'
APORTE_NUEVO_SALDO_COL = 'J'

# Columnas para totales (Las filas ahora son dinámicas)
COL_TOTAL_APORTES = 'H'
COL_GASTOS_ADMIN = 'K'
COL_TOTAL_GENERAL = 'K'

GASTO_POR_APORTE = 3000 
MAX_APORTES_PERMITIDOS = 6  # Nuevo límite establecido

def generar_recibo_solo_aportes(
    db_manager, 
    recibo_id: int,
    recibi_de_data: dict, 
    aportes_info: list = None,
    num_aportes_cobrables: int = None
):
    """
    Genera un recibo seleccionando la plantilla exacta según el número de aportes (1 a 6).
    Usa la fecha global HOY para permitir viajes en el tiempo.
    """
    if aportes_info is None:
        aportes_info = []

    # 1. Determinar cuántos aportes hay para elegir la plantilla
    num_aportes = len(aportes_info)

    # Validaciones de seguridad
    if num_aportes == 0:
        print("Error: No hay aportes para generar recibo.")
        return None
    
    if num_aportes > MAX_APORTES_PERMITIDOS:
        print(f"Advertencia: Hay {num_aportes} aportes. Se truncará a {MAX_APORTES_PERMITIDOS}.")
        aportes_info = aportes_info[:MAX_APORTES_PERMITIDOS]
        num_aportes = MAX_APORTES_PERMITIDOS

    try:
        os.makedirs(OUTPUT_FOLDER_PATH, exist_ok=True)
        
        # --- CAMBIO DE FECHA 1: Nombre del archivo (YYYYMMDD) ---
        file_name = f"Recibo_{recibo_id}_{get_hoy().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER_PATH, file_name)

        # 2. SELECCIÓN DE PLANTILLA SEGÚN CANTIDAD (Switch lógico)
        template_name = f"recibo_template_aporte{num_aportes}.xlsx"
        template_rel_path = os.path.join("templates", "recibo_template_aporte", template_name)
        template_abs_path = os.path.join(ASSETS_DIR, template_rel_path)

        if not os.path.exists(template_abs_path):
            print(f"❌ Error CRÍTICO: No se encontró la plantilla {template_name}")
            return None

        wb = load_workbook(template_abs_path)
        ws = wb.active

        # --- CABECERA ---
        ws[RECIBO_ID_CELL] = recibo_id
        
        # --- CAMBIO DE FECHA 2: Celda Excel (DD/MM/YYYY) ---
        ws[FECHA_CELL] = get_hoy().strftime("%d/%m/%Y")
        
        recibi_de_full_name = f"{recibi_de_data['nombres']} {recibi_de_data['apellidos']}".upper()
        ws[RECIBI_DE_CELL] = recibi_de_full_name
        ws[RECIBI_DE_CELL].alignment = Alignment(horizontal='center')

        # --- CUERPO (Llenar filas exactas) ---
        total_aportes_acumulado = 0
        
        for i in range(num_aportes):
            row_to_fill = APORTE_DATA_START_ROW + i
            
            socio_data, monto_aporte, saldo_socio_antes, saldo_socio_despues = aportes_info[i]
            
            formatted_socio_name = format_full_name_for_excel(
                socio_data['nombres'], 
                socio_data['apellidos'], 
                max_length=24 
            )
            
            # Escribir datos
            ws[f'{APORTE_NOMBRE_COL}{row_to_fill}'] = formatted_socio_name
            ws[f'{APORTE_SALDO_COL}{row_to_fill}'] = format_miles_colombian_int(saldo_socio_antes)
            ws[f'{APORTE_MONTO_COL}{row_to_fill}'] = format_miles_colombian_int(monto_aporte)
            ws[f'{APORTE_NUEVO_SALDO_COL}{row_to_fill}'] = format_miles_colombian_int(saldo_socio_despues)
            
            total_aportes_acumulado += monto_aporte

        # --- PIE DE PÁGINA (Cálculo dinámico de posiciones) ---
        row_total_aportes = APORTE_DATA_START_ROW + num_aportes
        row_gastos_admin = row_total_aportes + 2 
        row_total_general = row_total_aportes + 3

        # 1. Total Aportes
        ws[f'{COL_TOTAL_APORTES}{row_total_aportes}'] = format_miles_colombian_int(total_aportes_acumulado)

        # 2. Gastos Admin
        if num_aportes_cobrables is not None:
            cantidad_a_cobrar = num_aportes_cobrables
        else:
            cantidad_a_cobrar = num_aportes
            
        gastos_admin = GASTO_POR_APORTE * cantidad_a_cobrar
        ws[f'{COL_GASTOS_ADMIN}{row_gastos_admin}'] = format_miles_colombian_int(gastos_admin)
        
        # 3. Total General
        total_general = total_aportes_acumulado + gastos_admin
        ws[f'{COL_TOTAL_GENERAL}{row_total_general}'] = format_miles_colombian_int(total_general)

        # --- Guardar ---
        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"Error al generar recibo exacto de aportes: {e}")
        import traceback
        traceback.print_exc() 
        return None