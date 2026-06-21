import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment # Importa Alignment para centrar texto
from datetime import date
from config import (
    format_miles_colombian_int, 
    format_full_name_for_excel, 
    ASSETS_DIR, RECIBOS_OUTPUT_DIR, get_hoy, get_hoy_str
)

# --- Rutas y Constantes ---
# Construir rutas relativas a BASE_APP_DIR
TEMPLATE_RETIRO_REL_PATH = os.path.join("templates", "recibo_template_retiro.xlsx")
TEMPLATE_RETIRO_PATH = os.path.join(ASSETS_DIR, TEMPLATE_RETIRO_REL_PATH)

OUTPUT_FOLDER_PATH = RECIBOS_OUTPUT_DIR

# --- Constantes de Celda Específicas para recibo_template_retiro.xlsx ---
RECIBO_ID_CELL = 'B6'       # Numero de recibo
MONTO_RETIRA_CELL = 'F6'    # Monto que retira
FECHA_CELL = 'F8'           # Fecha
CONCEPTO_NOMBRES_CELL = 'C14' # DEVOLUCION PARCIAL DE APORTES DE {NOMBRES EN MAYUSCULAS}
APELLIDOS_SOCIO_CELL = 'C15' # APELLIDOS DEL SOCIO EN MAYUSCULAS
NOMBRE_COMPLETO_FIRMA_CELL = 'C18' # Nombre completo del socio en Mayusculas

# --- Funciones de Generación ---
def generar_recibo_retiro(
    recibo_id: int,
    socio_data: dict, # Diccionario con 'nombres', 'apellidos'
    monto_retiro: int
):
    """
    Genera un recibo de retiro/devolución.

    Args:
        recibo_id (int): El ID único del recibo.
        socio_data (dict): Diccionario con la información del socio, debe contener 'nombres' y 'apellidos'.
        monto_retiro (int): El monto total que se retira.
    """
    try:
        # Asegúrate de que la carpeta de salida exista
        os.makedirs(OUTPUT_FOLDER_PATH, exist_ok=True)

        file_name = f"Recibo_{recibo_id}_{get_hoy().strftime('%Y%m%d')}.xlsx"
        output_path = os.path.join(OUTPUT_FOLDER_PATH, file_name)

        wb = load_workbook(TEMPLATE_RETIRO_PATH)
        ws = wb.active

        # --- Reemplazar datos en el recibo ---
        ws[RECIBO_ID_CELL] = recibo_id
        ws[MONTO_RETIRA_CELL] = format_miles_colombian_int(monto_retiro)
        ws[FECHA_CELL] = get_hoy().strftime("%d/%m/%Y")

        # C14: "DEVOLUCION PARCIAL DE APORTES DE {nombres socio} solo los nombres y en mayusculas"
        nombres_socio_upper = socio_data['nombres'].upper()
        concepto_str = f"DEVOLUCION PARCIAL DE APORTES DE {nombres_socio_upper}"
        ws[CONCEPTO_NOMBRES_CELL] = concepto_str

        # C15: "los apellidos del socio en mayusculas"
        apellidos_socio_upper = socio_data['apellidos'].upper()
        ws[APELLIDOS_SOCIO_CELL] = apellidos_socio_upper

        # C18: "El nombre completo del socio en Mayusculas"
        nombre_completo_socio_upper = f"{socio_data['nombres']} {socio_data['apellidos']}".upper()
        ws[NOMBRE_COMPLETO_FIRMA_CELL] = nombre_completo_socio_upper
        # Opcional: Centrar el texto en C18 si es un campo de firma
        ws[NOMBRE_COMPLETO_FIRMA_CELL].alignment = Alignment(horizontal='center')

        # --- Guardar el recibo ---
        wb.save(output_path)
        return output_path

    except Exception as e:
        print(f"Error al generar recibo de retiro: {e}")
        import traceback
        traceback.print_exc()
        return None